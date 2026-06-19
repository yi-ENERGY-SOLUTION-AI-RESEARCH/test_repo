# -*- coding: utf-8 -*-
"""엑셀 양식 자동 작성 모듈.

추출된 경력 데이터를 종합 경력확인서 엑셀 양식에 맞게 변환하고 입력한다.

엑셀 열 구조 (3행이 헤더, 4행부터 데이터):
  B열: 성명
  C열: 생년월일
  D열: 등급
  E열: 기술자격
  F열: 학력
  G열: 근무처명 (근무기간) ← PDF 1페이지 근무처경력에서 추출
  H열: 착수일 (참여기간 시작일, 줄바꿈으로 구분)
  I열: 준공일 (참여일수) ← "YYYY.MM.DD (X일)" 형식, 줄바꿈 구분
  J열: 총근무기간 (실근무기간 일수)
  K열: 참여분야 (직무분야의 소분류: 변전/화력/송전/기타 등, 줄바꿈으로 구분)
  L열: 참여업무명 (줄바꿈으로 구분)
  M열: 발주처 (줄바꿈으로 구분)
  N열: 해당분야 (담당분야 고유값)
  O열: REMARKS
"""

import io
from datetime import date
from typing import Any

import openpyxl
from loguru import logger
from openpyxl.styles import Alignment


# 데이터가 시작되는 행 번호 (1행 제목, 2행 날짜, 3행 헤더, 4행부터 데이터)
_DATA_START_ROW = 4

# 참여분야 정규화 매핑: OCR이 추출하는 다양한 표현을 표준값으로 치환
_PARTICIPATION_FIELD_MAP: dict[str, str] = {
    "그 밖의 발송변배전시설": "변배전",
    "그 밖의 발전변배전시설": "변배전",
}


def _build_excel_row(career_info: dict[str, Any]) -> dict[str, str]:
    """경력 정보 딕셔너리를 엑셀 열별 값으로 변환.

    Args:
        career_info: process_pdf()가 반환한 경력 정보 딕셔너리.

    Returns:
        열 문자(A-O) → 셀 값 문자열 매핑 딕셔너리.
    """
    records: list[dict] = career_info.get("경력목록", [])

    # H열: 착수일 목록 (YYYY.MM.DD ~ 형식)
    start_dates = "\n".join(
        f"{r['착수일']} ~"
        for r in records
        if r.get("착수일")
    )

    # I열: 준공일(참여일수) 목록 (YYYY.MM.DD (X일) 형식)
    end_dates = "\n".join(
        f"{r['준공일']} ({r.get('참여일수', '')}일)"
        for r in records
        if r.get("준공일")
    )

    # J열: 총 실근무기간 (숫자일수 + "일")
    total_raw = career_info.get("실근무기간_일수", "")
    total_days = f"{total_raw}일" if total_raw else ""

    # K열: 참여분야 목록 (직무분야의 소분류, 예: 변전/화력/송전/기타 등)
    # _PARTICIPATION_FIELD_MAP을 통해 비표준 표현을 표준값으로 정규화한다.
    job_fields = "\n".join(
        _PARTICIPATION_FIELD_MAP.get(r.get("참여분야", ""), r.get("참여분야", ""))
        for r in records
    )

    # L열: 참여업무명 목록
    project_names = "\n".join(
        r.get("참여사업명", "") for r in records
    )

    # M열: 발주처 목록
    clients = "\n".join(
        r.get("발주자", "") for r in records
    )

    # N열: 해당분야 (담당분야 중 고유값만, 순서 유지)
    seen: set[str] = set()
    unique_fields: list[str] = []
    for r in records:
        field = r.get("담당분야", "")
        if field and field not in seen:
            seen.add(field)
            unique_fields.append(field)
    relevant_fields = "\n".join(unique_fields)

    return {
        "B": career_info.get("성명", ""),
        "C": career_info.get("생년월일", ""),
        "D": career_info.get("등급", ""),
        "E": career_info.get("기술자격", ""),
        "F": career_info.get("학력", ""),
        "G": career_info.get("근무처명_근무기간", ""),
        "H": start_dates,
        "I": end_dates,
        "J": total_days,
        "K": job_fields,
        "L": project_names,
        "M": clients,
        "N": relevant_fields,
        "O": "",
    }


def _find_next_empty_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """B열 기준으로 마지막 데이터 행 다음의 빈 행 번호를 반환.

    Args:
        ws: 대상 워크시트.

    Returns:
        데이터를 입력할 다음 빈 행 번호.
    """
    last_data_row = _DATA_START_ROW - 1
    for row_idx in range(_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value is not None:
            last_data_row = row_idx
    return last_data_row + 1


def _write_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_num: int,
    row_data: dict[str, str],
) -> None:
    """워크시트의 특정 행에 데이터를 입력하고 서식을 적용.

    Args:
        ws: 대상 워크시트.
        row_num: 데이터를 입력할 행 번호.
        row_data: 열 문자 → 값 딕셔너리.
    """
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for col_letter, value in row_data.items():
        cell = ws[f"{col_letter}{row_num}"]
        cell.value = value if value else None
        # 줄바꿈 포함 셀은 자동 줄 바꿈 서식 적용
        if value and "\n" in value:
            cell.alignment = wrap_align


def fill_excel_template(
    template_bytes: bytes,
    all_career_data: list[dict[str, Any]],
    append_mode: bool = False,
) -> bytes:
    """엑셀 템플릿에 경력 데이터를 채워 반환.

    Args:
        template_bytes: 종합 경력확인서 엑셀 템플릿 파일 바이트.
        all_career_data: 경력 데이터 목록 (각 항목이 한 행).
        append_mode: True이면 기존 데이터 뒤에 추가, False이면 4행부터 덮어쓰기.

    Returns:
        데이터가 입력된 엑셀 파일 바이트.
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    ws["O2"] = date.today().strftime("%Y-%m-%d")

    if append_mode:
        start_row = _find_next_empty_row(ws)
    else:
        start_row = _DATA_START_ROW

    for idx, career_info in enumerate(all_career_data):
        row_num = start_row + idx
        row_data = _build_excel_row(career_info)
        _write_row(ws, row_num, row_data)

        name = career_info.get("성명", f"항목{idx + 1}")
        record_count = len(career_info.get("경력목록", []))
        logger.info(
            f"✍️  {row_num}행 입력 완료: {name} (경력 {record_count}건)"
        )

    output = io.BytesIO()
    wb.save(output)
    logger.info(f"📊 엑셀 파일 생성 완료 (총 {len(all_career_data)}행 입력)")
    return output.getvalue()

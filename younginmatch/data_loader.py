# -*- coding: utf-8 -*-
"""CV 및 회사 실적 데이터 로더.

CV.xlsx(경력확인서)와 영인에너지솔루션_실적정리.xlsx를 읽어
GPT 분석에 사용할 구조화된 딕셔너리로 변환한다.
"""

from pathlib import Path

import openpyxl
from loguru import logger

# 데이터 파일 경로 (data_loader.py 기준 동일 폴더)
_BASE_DIR = Path(__file__).parent
CV_FILE = _BASE_DIR / "CV.xlsx"
SILJEOK_FILE = _BASE_DIR / "영인에너지솔루션_실적정리.xlsx"

# CV.xlsx 열 인덱스 매핑 (B=2 ~ O=15, 1-indexed)
_CV_COLUMNS = {
    "성명": 2,
    "생년월일": 3,
    "등급": 4,
    "기술자격": 5,
    "학력": 6,
    "근무처명_근무기간": 7,
    "착수일": 8,
    "준공일_참여일수": 9,
    "총근무기간_일수": 10,
    "참여분야": 11,
    "참여업무명": 12,
    "발주처": 13,
    "해당분야": 14,
    "remarks": 15,
}

# CV 데이터 시작 행 (1행 제목, 2행 날짜, 3행 헤더, 4행부터 데이터)
_CV_DATA_START_ROW = 4


def _cell_to_lines(value: object) -> list[str]:
    """셀 값을 줄바꿈 기준으로 분리한 리스트로 변환.

    Args:
        value: 셀 값 (문자열, 숫자, None 등).

    Returns:
        비어있지 않은 줄 문자열 리스트. 값이 None이면 빈 리스트.
    """
    if value is None:
        return []
    text = str(value).strip()
    return [line.strip() for line in text.split("\n") if line.strip()]


def _days_to_years(days_str: str | None) -> float:
    """총근무기간(일수 문자열)을 경력년수(float)로 변환.

    Args:
        days_str: "3774일" 또는 "3774" 형태의 문자열.

    Returns:
        소수점 1자리 경력년수. 변환 불가 시 0.0.
    """
    if not days_str:
        return 0.0
    digits = "".join(c for c in str(days_str) if c.isdigit())
    if not digits:
        return 0.0
    return round(int(digits) / 365, 1)


def load_cv_data() -> list[dict]:
    """CV.xlsx에서 인력 데이터를 읽어 딕셔너리 리스트로 반환.

    Returns:
        인력별 딕셔너리 리스트. 각 딕셔너리는 다음 키를 포함:
        - 성명 (str)
        - 생년월일 (str)
        - 등급 (str): 특급기술자 / 고급기술자 등
        - 기술자격 (list[str])
        - 학력 (list[str])
        - 근무처명_근무기간 (list[str])
        - 착수일 (list[str])
        - 준공일_참여일수 (list[str])
        - 총근무기간_일수 (str)
        - 경력년수 (float): 총근무기간을 연수로 환산
        - 참여분야 (list[str])
        - 참여업무명 (list[str])
        - 발주처 (list[str])
        - 해당분야 (str)
        - remarks (str)
        - 프로젝트수 (int): 참여업무명 건수
    """
    if not CV_FILE.exists():
        logger.error(f"CV 파일을 찾을 수 없습니다: {CV_FILE}")
        return []

    wb = openpyxl.load_workbook(CV_FILE)
    ws = wb.active
    results: list[dict] = []

    for row_idx in range(_CV_DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row_idx, _CV_COLUMNS["성명"]).value
        if not name:
            continue

        raw_days = ws.cell(row_idx, _CV_COLUMNS["총근무기간_일수"]).value
        days_str = str(raw_days).strip() if raw_days else ""

        person: dict = {
            "성명": str(name).strip(),
            "생년월일": str(ws.cell(row_idx, _CV_COLUMNS["생년월일"]).value or "").strip(),
            "등급": str(ws.cell(row_idx, _CV_COLUMNS["등급"]).value or "").strip(),
            "기술자격": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["기술자격"]).value),
            "학력": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["학력"]).value),
            "근무처명_근무기간": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["근무처명_근무기간"]).value),
            "착수일": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["착수일"]).value),
            "준공일_참여일수": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["준공일_참여일수"]).value),
            "총근무기간_일수": days_str,
            "경력년수": _days_to_years(days_str),
            "참여분야": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["참여분야"]).value),
            "참여업무명": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["참여업무명"]).value),
            "발주처": _cell_to_lines(ws.cell(row_idx, _CV_COLUMNS["발주처"]).value),
            "해당분야": str(ws.cell(row_idx, _CV_COLUMNS["해당분야"]).value or "").strip(),
            "remarks": str(ws.cell(row_idx, _CV_COLUMNS["remarks"]).value or "").strip(),
        }
        person["프로젝트수"] = len(person["참여업무명"])
        results.append(person)
        logger.info(f"✅ CV 로드: {person['성명']} ({person['등급']}, {person['경력년수']}년, {person['프로젝트수']}건)")

    logger.info(f"📋 총 {len(results)}명 인력 데이터 로드 완료")
    return results


def load_siljeok_data() -> dict[str, list[dict]]:
    """영인에너지솔루션_실적정리.xlsx에서 회사 실적을 읽어 시트별 딕셔너리로 반환.

    Returns:
        시트명 → 프로젝트 딕셔너리 리스트 매핑.
        각 프로젝트 딕셔너리: {"발주처": str, "프로젝트명": str, "수행기간": str}
    """
    if not SILJEOK_FILE.exists():
        logger.error(f"실적 파일을 찾을 수 없습니다: {SILJEOK_FILE}")
        return {}

    wb = openpyxl.load_workbook(SILJEOK_FILE)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        projects: list[dict] = []

        # 2행이 헤더(발주처/프로젝트명/수행기간), 3행부터 데이터
        for row_idx in range(3, ws.max_row + 1):
            client = ws.cell(row_idx, 1).value
            project = ws.cell(row_idx, 2).value
            period = ws.cell(row_idx, 3).value
            if not client and not project:
                continue
            projects.append({
                "발주처": str(client or "").strip(),
                "프로젝트명": str(project or "").strip(),
                "수행기간": str(period or "").strip(),
            })

        result[sheet_name] = projects
        logger.info(f"📁 실적 시트 [{sheet_name}]: {len(projects)}건")

    total = sum(len(v) for v in result.values())
    logger.info(f"🏢 총 {total}건 회사 실적 로드 완료")
    return result


def build_context_text(cv_list: list[dict], siljeok: dict[str, list[dict]]) -> str:
    """CV 및 실적 데이터를 GPT에 전달할 텍스트 컨텍스트로 변환.

    Args:
        cv_list: load_cv_data() 반환값.
        siljeok: load_siljeok_data() 반환값.

    Returns:
        GPT 프롬프트에 삽입할 구조화된 텍스트 문자열.
    """
    lines: list[str] = ["=== 참여 가능 인력 목록 ===\n"]

    for i, p in enumerate(cv_list, 1):
        lines.append(f"[인력 {i}] {p['성명']}")
        lines.append(f"  - 등급: {p['등급']}")
        lines.append(f"  - 기술자격: {', '.join(p['기술자격'])}")
        lines.append(f"  - 학력: {', '.join(p['학력'])}")
        lines.append(f"  - 경력년수: {p['경력년수']}년 (총 {p['총근무기간_일수']})")
        lines.append(f"  - 해당분야: {p['해당분야']}")
        lines.append(f"  - 참여분야: {', '.join(set(p['참여분야']))}")
        lines.append(f"  - 참여업무명 ({p['프로젝트수']}건):")
        for j, proj_name in enumerate(p["참여업무명"], 1):
            발주처 = p["발주처"][j - 1] if j - 1 < len(p["발주처"]) else ""
            lines.append(f"    {j}. {proj_name} / 발주처: {발주처}")
        lines.append("")

    lines.append("=== 회사 실적 현황 ===\n")
    for sheet_name, projects in siljeok.items():
        lines.append(f"[{sheet_name}] ({len(projects)}건)")
        for proj in projects:
            lines.append(f"  - {proj['프로젝트명']} | 발주처: {proj['발주처']} | 기간: {proj['수행기간']}")
        lines.append("")

    return "\n".join(lines)
